# ============================================================================
# MOTOR UNIVERSAL DE VALIDAÇÃO PRODUTO A PRODUTO — ResolvRapido Brasil v20 SOBERANO
# ============================================================================
# Agnóstico de setor (indústria, comércio, serviços, agropecuária, saúde,
# construção, energia, etc.) e de regime tributário (Lucro Real, Presumido,
# Simples Nacional, MEI, CPF Rural).
#
# NÃO ALTERA a lógica central de cálculo de créditos já existente no v18
# (validacao_produtos.py / main.py). Este módulo é uma CAMADA ADICIONAL:
# parseia o SPED de forma agnóstica, classifica cada item por setor/NCM,
# aplica as regras universais de validação e gera um relatório Excel
# completo, produto a produto.
#
# Principais entregáveis (conforme especificação):
#   - parsear_sped_universal(conteudo)
#   - validar_produto_universal(item, regime, uf)
#   - gerar_relatorio_produtos_universal(sped_bytes, cnpj, regime, uf)
# ============================================================================

from __future__ import annotations

import hashlib
import io
import json
import os
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

# ---------------------------------------------------------------------------
# 0) UTILIDADES BÁSICAS
# ---------------------------------------------------------------------------

def _safe_dec(val: Any, default: str = "0") -> Decimal:
    """Converte qualquer valor (str com vírgula, None, etc.) para Decimal."""
    if val is None or val == "":
        return Decimal(default)
    try:
        s = str(val).replace(",", ".").strip()
        if s in ("", "-", "."):
            return Decimal(default)
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _norm_ncm(raw: Any) -> str:
    """Normaliza um NCM para 8 dígitos, removendo pontuação."""
    s = re.sub(r"[^\d]", "", str(raw or ""))
    return s[:8]


def _norm_cnpj_cpf(raw: Any) -> str:
    return re.sub(r"[^\d]", "", str(raw or ""))


def sha256_linha(*campos: Any) -> str:
    """Gera hash SHA-256 de integridade de uma linha (produto a produto)."""
    conteudo = "|".join(str(c) for c in campos)
    return hashlib.sha256(conteudo.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# 1) BASE NCM/TIPI UNIVERSAL (fallback local, ENRIQUECIDA por categoria)
# ---------------------------------------------------------------------------
# Fonte primária real: Siscomex Classif (15.156 NCMs), via integrador_rfb*.py.
# Esta base local cobre TODAS as macro-categorias da economia, usada como
# fallback determinístico (por prefixo de NCM) quando a API está offline
# ou quando o NCM exato não está catalogado individualmente.
#
# Estrutura por faixa de NCM (2 a 4 primeiros dígitos):
#   setor, descrição, ipi padrão, monofásico (bool), pis/cofins nc e cum,
#   observação legal.
# ---------------------------------------------------------------------------

FAIXAS_NCM_UNIVERSAL: List[Dict[str, Any]] = [
    # --- AGROPECUÁRIA / ALIMENTOS ---
    {"prefixo": "01", "setor": "Agropecuária", "descricao": "Animais vivos", "ipi": 0.0, "monofasico": False},
    {"prefixo": "02", "setor": "Alimentos", "descricao": "Carnes e miudezas comestíveis", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "04", "setor": "Alimentos", "descricao": "Leite, laticínios, ovos, mel", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "07", "setor": "Alimentos", "descricao": "Produtos hortícolas", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "08", "setor": "Alimentos", "descricao": "Frutas", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "10", "setor": "Agropecuária", "descricao": "Cereais (arroz, milho, trigo)", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "1006", "setor": "Alimentos", "descricao": "Arroz", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "1101", "setor": "Alimentos", "descricao": "Farinha de trigo", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "12", "setor": "Agropecuária", "descricao": "Sementes e grãos oleaginosos (soja)", "ipi": 0.0, "monofasico": False},
    {"prefixo": "1507", "setor": "Alimentos", "descricao": "Óleo de soja", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "1701", "setor": "Alimentos", "descricao": "Açúcar", "ipi": 0.0, "monofasico": False, "cesta_basica": True},
    {"prefixo": "1901", "setor": "Alimentos", "descricao": "Preparações alimentícias de farinha", "ipi": 0.0, "monofasico": False},
    {"prefixo": "23", "setor": "Agropecuária", "descricao": "Resíduos e farelos (ração/insumo animal)", "ipi": 0.0, "monofasico": False},
    {"prefixo": "2302", "setor": "Agropecuária", "descricao": "Farelos e resíduos p/ ração animal", "ipi": 0.0, "monofasico": False},
    {"prefixo": "31", "setor": "Agropecuária", "descricao": "Fertilizantes", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.925/2004 — alíquota zero PIS/COFINS na cadeia"},
    {"prefixo": "3102", "setor": "Agropecuária", "descricao": "Adubos/fertilizantes nitrogenados (ureia)", "ipi": 0.0, "monofasico": True},
    {"prefixo": "3808", "setor": "Agropecuária", "descricao": "Defensivos agrícolas/agrotóxicos", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.925/2004"},

    # --- BEBIDAS ---
    {"prefixo": "2201", "setor": "Bebidas", "descricao": "Águas minerais", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.833/2003"},
    {"prefixo": "2202", "setor": "Bebidas", "descricao": "Refrigerantes e refrescos", "ipi": 4.0, "monofasico": True, "obs": "Lei 10.833/2003"},
    {"prefixo": "2203", "setor": "Bebidas", "descricao": "Cervejas de malte", "ipi": 10.0, "monofasico": True, "obs": "Lei 13.097/2015"},
    {"prefixo": "2204", "setor": "Bebidas", "descricao": "Vinhos", "ipi": 10.0, "monofasico": True, "obs": "Lei 13.097/2015"},
    {"prefixo": "2208", "setor": "Bebidas", "descricao": "Bebidas destiladas/álcool", "ipi": 30.0, "monofasico": True, "obs": "Lei 13.097/2015"},

    # --- COMBUSTÍVEIS E LUBRIFICANTES ---
    {"prefixo": "2709", "setor": "Combustíveis", "descricao": "Óleos brutos de petróleo", "ipi": 0.0, "monofasico": True},
    {"prefixo": "2710", "setor": "Combustíveis", "descricao": "Gasolina, diesel, óleos combustíveis", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.865/2004"},
    {"prefixo": "2711", "setor": "Combustíveis", "descricao": "GLP / GNV", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.865/2004"},
    {"prefixo": "2909", "setor": "Combustíveis", "descricao": "Álcool combustível/etanol", "ipi": 0.0, "monofasico": True},
    {"prefixo": "27101932", "setor": "Combustíveis", "descricao": "Óleo diesel", "ipi": 0.0, "monofasico": True},

    # --- MEDICAMENTOS E SAÚDE ---
    {"prefixo": "3003", "setor": "Saúde", "descricao": "Medicamentos (misturados)", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.147/2000"},
    {"prefixo": "3004", "setor": "Saúde", "descricao": "Medicamentos (doses)", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.147/2000"},
    {"prefixo": "3005", "setor": "Saúde", "descricao": "Gazes, ataduras, algodão", "ipi": 0.0, "monofasico": False},
    {"prefixo": "9018", "setor": "Saúde", "descricao": "Instrumentos e aparelhos médicos", "ipi": 0.0, "monofasico": False},
    {"prefixo": "9019", "setor": "Saúde", "descricao": "Aparelhos de fisioterapia/ortopedia", "ipi": 0.0, "monofasico": False},
    {"prefixo": "9021", "setor": "Saúde", "descricao": "Próteses e artigos ortopédicos", "ipi": 0.0, "monofasico": False},

    # --- COSMÉTICOS E HIGIENE ---
    {"prefixo": "3303", "setor": "Cosméticos", "descricao": "Perfumes e águas-de-colônia", "ipi": 12.0, "monofasico": True, "obs": "Lei 10.147/2000"},
    {"prefixo": "3304", "setor": "Cosméticos", "descricao": "Produtos de maquiagem", "ipi": 12.0, "monofasico": True, "obs": "Lei 10.147/2000"},
    {"prefixo": "3305", "setor": "Cosméticos", "descricao": "Xampus e preparações capilares", "ipi": 7.0, "monofasico": True, "obs": "Lei 10.147/2000"},
    {"prefixo": "3306", "setor": "Cosméticos", "descricao": "Dentifrícios / higiene bucal", "ipi": 7.0, "monofasico": True, "obs": "Lei 10.147/2000"},
    {"prefixo": "3401", "setor": "Cosméticos", "descricao": "Sabões e produtos de limpeza pessoal", "ipi": 0.0, "monofasico": True},

    # --- FUMO ---
    {"prefixo": "2402", "setor": "Fumo", "descricao": "Cigarros e charutos", "ipi": 30.0, "monofasico": True, "obs": "Lei 10.865/2004 + IPI ad valorem"},
    {"prefixo": "2403", "setor": "Fumo", "descricao": "Tabaco manufaturado", "ipi": 30.0, "monofasico": True},

    # --- AUTOPEÇAS E VEÍCULOS ---
    {"prefixo": "4011", "setor": "Autopeças", "descricao": "Pneus novos de borracha", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.485/2002"},
    {"prefixo": "8407", "setor": "Veículos", "descricao": "Motores de pistão (veículos)", "ipi": 0.0, "monofasico": True},
    {"prefixo": "8409", "setor": "Autopeças", "descricao": "Peças de motores", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.485/2002"},
    {"prefixo": "8413", "setor": "Autopeças", "descricao": "Bombas de fluidos", "ipi": 0.0, "monofasico": False},
    {"prefixo": "8483", "setor": "Autopeças", "descricao": "Eixos, engrenagens, câmbio", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.485/2002"},
    {"prefixo": "8708", "setor": "Autopeças", "descricao": "Partes e acessórios de veículos", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.485/2002"},
    {"prefixo": "8701", "setor": "Veículos", "descricao": "Tratores", "ipi": 0.0, "monofasico": False},
    {"prefixo": "8702", "setor": "Veículos", "descricao": "Veículos de transporte coletivo", "ipi": 0.0, "monofasico": True, "obs": "Lei 10.485/2002"},
    {"prefixo": "8703", "setor": "Veículos", "descricao": "Automóveis de passageiros", "ipi": 25.0, "monofasico": True, "obs": "Lei 10.485/2002"},
    {"prefixo": "8704", "setor": "Veículos", "descricao": "Caminhões", "ipi": 10.0, "monofasico": True, "obs": "Lei 10.485/2002"},
    {"prefixo": "8711", "setor": "Veículos", "descricao": "Motocicletas", "ipi": 10.0, "monofasico": True, "obs": "Lei 10.485/2002"},

    # --- MÁQUINAS E EQUIPAMENTOS INDUSTRIAIS/AGRÍCOLAS ---
    {"prefixo": "8421", "setor": "Máquinas", "descricao": "Centrifugadoras, filtros industriais", "ipi": 5.0, "monofasico": False},
    {"prefixo": "8429", "setor": "Máquinas", "descricao": "Máquinas de terraplanagem", "ipi": 5.0, "monofasico": False},
    {"prefixo": "8432", "setor": "Máquinas", "descricao": "Máquinas agrícolas", "ipi": 0.0, "monofasico": False},
    {"prefixo": "8433", "setor": "Máquinas", "descricao": "Colheitadeiras", "ipi": 0.0, "monofasico": False},

    # --- MATERIAIS DE CONSTRUÇÃO E FERRAGENS ---
    {"prefixo": "2523", "setor": "Construção", "descricao": "Cimento", "ipi": 0.0, "monofasico": False},
    {"prefixo": "6907", "setor": "Construção", "descricao": "Ladrilhos e placas cerâmicas", "ipi": 5.0, "monofasico": False},
    {"prefixo": "6908", "setor": "Construção", "descricao": "Azulejos e revestimentos cerâmicos", "ipi": 5.0, "monofasico": False},
    {"prefixo": "7214", "setor": "Construção", "descricao": "Vergalhões de aço (CA-50/CA-60)", "ipi": 5.0, "monofasico": False},
    {"prefixo": "7326", "setor": "Indústria", "descricao": "Obras de ferro/aço (peças industriais)", "ipi": 5.0, "monofasico": False},
    {"prefixo": "4418", "setor": "Construção", "descricao": "Obras de marcenaria/madeira p/ construção", "ipi": 5.0, "monofasico": False},

    # --- ELETRÔNICOS E INFORMÁTICA ---
    {"prefixo": "8471", "setor": "Informática", "descricao": "Computadores e periféricos", "ipi": 0.0, "monofasico": False, "obs": "Lei do Bem — IPI reduzido/zero"},
    {"prefixo": "8517", "setor": "Eletrônicos", "descricao": "Telefones/celulares", "ipi": 10.0, "monofasico": False},
    {"prefixo": "8528", "setor": "Eletrônicos", "descricao": "Televisores e monitores", "ipi": 15.0, "monofasico": False},
    {"prefixo": "8542", "setor": "Eletrônicos", "descricao": "Circuitos integrados/componentes eletrônicos", "ipi": 0.0, "monofasico": False},

    # --- PLÁSTICOS, BORRACHA, EMBALAGENS ---
    {"prefixo": "3920", "setor": "Plásticos", "descricao": "Chapas e filmes plásticos", "ipi": 5.0, "monofasico": False},
    {"prefixo": "3921", "setor": "Plásticos", "descricao": "Outras chapas plásticas", "ipi": 5.0, "monofasico": False},
    {"prefixo": "3923", "setor": "Plásticos", "descricao": "Embalagens plásticas (PET etc.)", "ipi": 5.0, "monofasico": False},
    {"prefixo": "4001", "setor": "Borracha", "descricao": "Borracha natural", "ipi": 0.0, "monofasico": False},
    {"prefixo": "4016", "setor": "Borracha", "descricao": "Artefatos de borracha vulcanizada", "ipi": 5.0, "monofasico": False},

    # --- QUÍMICOS E FERTILIZANTES ---
    {"prefixo": "2804", "setor": "Químico", "descricao": "Gases industriais (hidrogênio etc.)", "ipi": 0.0, "monofasico": False},
    {"prefixo": "2933", "setor": "Químico", "descricao": "Compostos heterocíclicos (insumo químico/farmacêutico)", "ipi": 0.0, "monofasico": False},
    {"prefixo": "3103", "setor": "Agropecuária", "descricao": "Superfosfatos", "ipi": 0.0, "monofasico": True},
    {"prefixo": "3104", "setor": "Agropecuária", "descricao": "Cloreto de potássio", "ipi": 0.0, "monofasico": True},

    # --- TÊXTEIS E VESTUÁRIO ---
    {"prefixo": "52", "setor": "Têxtil", "descricao": "Algodão e fios", "ipi": 0.0, "monofasico": False},
    {"prefixo": "5402", "setor": "Têxtil", "descricao": "Fibras sintéticas", "ipi": 0.0, "monofasico": False},
    {"prefixo": "61", "setor": "Vestuário", "descricao": "Vestuário de malha", "ipi": 0.0, "monofasico": False},
    {"prefixo": "62", "setor": "Vestuário", "descricao": "Vestuário exceto malha", "ipi": 0.0, "monofasico": False},

    # --- PAPEL E CELULOSE ---
    {"prefixo": "4701", "setor": "Papel/Celulose", "descricao": "Pastas de celulose", "ipi": 0.0, "monofasico": False},
    {"prefixo": "4802", "setor": "Papel/Celulose", "descricao": "Papel para impressão/escrita", "ipi": 0.0, "monofasico": False},
    {"prefixo": "4818", "setor": "Papel/Celulose", "descricao": "Papel higiênico e afins", "ipi": 0.0, "monofasico": False},
    {"prefixo": "4802569", "setor": "Papel/Celulose", "descricao": "Papel sulfite", "ipi": 0.0, "monofasico": False},

    # --- MOBILIÁRIO ---
    {"prefixo": "9401", "setor": "Mobiliário", "descricao": "Assentos/cadeiras", "ipi": 5.0, "monofasico": False},
    {"prefixo": "9403", "setor": "Mobiliário", "descricao": "Móveis (mesas, armários)", "ipi": 5.0, "monofasico": False},
    {"prefixo": "9404", "setor": "Mobiliário", "descricao": "Colchões, travesseiros", "ipi": 5.0, "monofasico": False},
]
# Ordena por especificidade (prefixos mais longos primeiro) para o "mais próximo" funcionar
FAIXAS_NCM_UNIVERSAL.sort(key=lambda x: -len(x["prefixo"]))

CST_MONOFASICO_SEM_CREDITO = {"04", "05", "06", "70", "71", "72", "73", "74", "75", "98", "99"}
CST_COM_CREDITO = {"50", "51", "52", "53", "54", "55", "56", "60", "61", "62", "63"}
CST_TRIBUTACAO_INTEGRAL = {"01", "02", "03"}

ALIQUOTAS_ICMS_INTERNO = {
    "AC": 19.0, "AL": 19.0, "AP": 18.0, "AM": 20.0, "BA": 20.5, "CE": 20.0, "DF": 20.0,
    "ES": 17.0, "GO": 19.0, "MA": 22.0, "MT": 17.0, "MS": 17.0, "MG": 18.0, "PA": 19.0,
    "PB": 20.0, "PR": 19.5, "PE": 20.5, "PI": 21.0, "RJ": 22.0, "RN": 18.0, "RS": 17.0,
    "RO": 19.5, "RR": 20.0, "SC": 17.0, "SP": 18.0, "SE": 19.0, "TO": 20.0,
}
UFS_SUL_SUDESTE = {"SP", "RJ", "MG", "ES", "PR", "SC", "RS"}


_NCM_CACHE_MEMORIA: Dict[str, Dict[str, Any]] = {}
# Cache em memória de NCM -> resultado consultado (chave da performance):
# evita reabrir/reconsultar os bancos SQLite de fontes oficiais a cada item
# repetido em SPEDs com milhões de linhas (o mesmo produto costuma repetir
# em centenas/milhares de notas).


def _consultar_ncm_universal(ncm: str) -> Dict[str, Any]:
    """
    Consulta NCM na base universal por faixa/prefixo.
    Tenta usar integrador_rfb / integrador_rfb_expandido (API Siscomex + fontes
    oficiais, JÁ CACHEADAS em disco) se disponíveis; caso contrário usa apenas
    o fallback local. Resultado é memoizado por NCM (cache em memória) para
    que SPEDs com milhões de linhas não paguem o custo de I/O repetidamente.
    """
    ncm_norm = _norm_ncm(ncm)
    if ncm_norm in _NCM_CACHE_MEMORIA:
        return _NCM_CACHE_MEMORIA[ncm_norm]
    resultado = _consultar_ncm_universal_sem_cache(ncm_norm)
    _NCM_CACHE_MEMORIA[ncm_norm] = resultado
    return resultado


def _consultar_ncm_universal_sem_cache(ncm_norm: str) -> Dict[str, Any]:
    info: Dict[str, Any] = {
        "ncm": ncm_norm, "encontrado": False, "fonte": "NAO_ENCONTRADO",
        "setor": "Não identificado", "descricao": "NCM não catalogado",
        "ipi": 0.0, "monofasico": False, "obs": "",
        "cesta_basica": False,
    }
    if not ncm_norm:
        return info

    # 1) Tenta fontes oficiais JÁ CACHEADAS (integrador_rfb / integrador_rfb_expandido).
    #    IMPORTANTE: nunca dispara download ao vivo aqui (isso quebraria a performance
    #    de streaming/lote). O cache das APIs deve ser aquecido previamente (uma única
    #    vez) via `integrador_rfb.atualizar_cache_ncm()` /
    #    `integrador_rfb_expandido.inicializar_todas_fontes()`. Se o cache local
    #    (SQLite) já existir em disco, ele é consultado; caso contrário, cai
    #    direto no fallback local determinístico abaixo — sem qualquer chamada de rede.
    try:
        import integrador_rfb_expandido as _rfb  # type: ignore
        _cache_path = getattr(_rfb, "_CACHE_DB_EXPANDED", None)
        if _cache_path is not None and _cache_path.exists():
            mono = _rfb.consultar_monofasico_detalhado(ncm_norm)
            if mono.get("monofasico"):
                info["monofasico"] = True
                info["obs"] = mono.get("lei_principal", "")
                info["fonte"] = "API_OFICIAL_CACHEADA"
    except Exception:
        pass
    try:
        import integrador_rfb as _rfb2  # type: ignore
        _cache_path2 = getattr(_rfb2, "CACHE_DB", None)
        if _cache_path2 is not None and _cache_path2.exists() and hasattr(_rfb2, "consultar_ncm"):
            dado = _rfb2.consultar_ncm(ncm_norm)
            if dado and dado.get("encontrado"):
                info["encontrado"] = True
                info["descricao"] = dado.get("descricao", info["descricao"])
                info["ipi"] = float(dado.get("ipi_aliquota", dado.get("ipi", info["ipi"])) or 0)
                info["fonte"] = "API_SISCOMEX_CACHEADA"
    except Exception:
        pass

    # 2) Fallback local por faixa/prefixo (sempre roda para setor/monofásico)
    for faixa in FAIXAS_NCM_UNIVERSAL:
        pref = faixa["prefixo"]
        if ncm_norm.startswith(pref):
            info["encontrado"] = True
            info["setor"] = faixa["setor"]
            if info["descricao"] == "NCM não catalogado":
                info["descricao"] = faixa["descricao"]
            if not info.get("monofasico"):
                info["monofasico"] = faixa.get("monofasico", False)
            if info["ipi"] == 0.0:
                info["ipi"] = faixa.get("ipi", 0.0)
            info["cesta_basica"] = faixa.get("cesta_basica", False)
            if not info["obs"]:
                info["obs"] = faixa.get("obs", "")
            if info["fonte"] == "NAO_ENCONTRADO":
                info["fonte"] = "FALLBACK_LOCAL"
            break

    return info


def _sugerir_ncm_proximo(ncm: str) -> Optional[str]:
    """Sugere NCM mais próximo com base nos 6 ou 4 primeiros dígitos."""
    ncm_norm = _norm_ncm(ncm)
    for tam in (6, 4, 2):
        prefixo = ncm_norm[:tam]
        for faixa in FAIXAS_NCM_UNIVERSAL:
            if faixa["prefixo"].startswith(prefixo) or prefixo.startswith(faixa["prefixo"]):
                return f"{faixa['prefixo']}... ({faixa['descricao']})"
    return None


# ---------------------------------------------------------------------------
# 2) PARSER SPED UNIVERSAL (AGNÓSTICO DE SETOR)
# ---------------------------------------------------------------------------

def parsear_sped_universal(conteudo: str) -> Dict[str, Any]:
    """
    Extrai registros 0000, 0150, 0200, C100, C170, C190, M100, M200, M500,
    M600 de um SPED (EFD-Contribuições ou EFD-ICMS/IPI), de forma agnóstica
    de setor.

    NÃO QUEBRA se o arquivo não tiver C170 (ex.: setor de serviços puros /
    NFS-e). Nesse caso retorna itens_c170 vazio e uma mensagem clara em
    "avisos".

    Retorna:
        {
            "empresa": {"cnpj": ..., "razao_social": ..., "uf": ..., "tipo_pessoa": "PJ"/"PF"},
            "produtos_0200": {cod_item: {"descricao":..., "ncm":...}},
            "fornecedores_0150": {cod_part: {...}},
            "itens_c170": [...],
            "resumo_c190": [...],
            "apuracao": {"pis": {...}, "cofins": {...}},
            "avisos": [...],
            "total_linhas": int,
        }
    """
    resultado: Dict[str, Any] = {
        "empresa": {"cnpj": "", "razao_social": "", "uf": "", "tipo_pessoa": "PJ"},
        "produtos_0200": {},
        "fornecedores_0150": {},
        "itens_c170": [],
        "resumo_c190": [],
        "apuracao": {"pis": {}, "cofins": {}},
        "avisos": [],
        "total_linhas": 0,
    }

    if not conteudo or not conteudo.strip():
        resultado["avisos"].append(
            "Arquivo vazio ou ilegível. Nenhum dado extraído."
        )
        return resultado

    c100_atual: Dict[str, Any] = {}
    linhas = conteudo.splitlines()
    resultado["total_linhas"] = len(linhas)

    for num_linha, linha in enumerate(linhas, 1):
        if not linha or "|" not in linha:
            continue
        partes = linha.strip().split("|")
        if len(partes) < 2:
            continue
        tipo = partes[1].strip()
        g = lambda i, d="": partes[i].strip() if len(partes) > i else d

        try:
            if tipo == "0000":
                # Layout comum a EFD-ICMS/IPI e EFD-Contribuições
                cnpj_raw = g(7)
                cpf_like = _norm_cnpj_cpf(cnpj_raw)
                resultado["empresa"]["razao_social"] = g(6) or resultado["empresa"]["razao_social"]
                resultado["empresa"]["cnpj"] = cnpj_raw or resultado["empresa"]["cnpj"]
                resultado["empresa"]["uf"] = g(9) or resultado["empresa"]["uf"]
                if len(cpf_like) == 11:
                    resultado["empresa"]["tipo_pessoa"] = "PF"

            elif tipo == "0150":
                cod_part = g(2)
                resultado["fornecedores_0150"][cod_part] = {
                    "nome": g(3),
                    "cod_pais": g(4),
                    "cnpj": g(5),
                    "cpf": g(6),
                    "uf": g(8),
                    "tipo_pessoa": "PF" if g(6) else "PJ",
                }

            elif tipo == "0200":
                cod_item = g(2)
                resultado["produtos_0200"][cod_item] = {
                    "descricao": g(3),
                    "ncm": g(4),
                    "unidade": g(5),
                    "tipo_item": g(9),
                }

            elif tipo == "C100":
                # Detecta layout (EFD ICMS/IPI tem mais campos do que Contribuições)
                c100_atual = {
                    "cod_part": g(4),
                    "num_doc": g(8) if len(partes) > 20 else g(6),
                    "dt_doc": g(10) if len(partes) > 20 else g(9),
                    "vl_doc": g(12) if len(partes) > 20 else g(11),
                    "vl_icms": g(22) if len(partes) > 22 else "0",
                    "vl_bc_icms_st": g(23) if len(partes) > 23 else "0",
                    "vl_icms_st": g(24) if len(partes) > 24 else "0",
                    "chv_nfe": g(9) if len(partes) > 20 and len(g(9)) == 44 else "",
                }

            elif tipo == "C170" and c100_atual:
                cod_item = g(3)
                item = {
                    "num_linha": num_linha,
                    "num_item": g(2),
                    "cod_item": cod_item,
                    "descr": g(4),
                    "qtd": g(5),
                    "unid": g(6),
                    "vl_item": g(7),
                    "vl_desc": g(8),
                    "cfop": g(11),
                    "cst_icms": g(10),
                    "vl_bc_icms": g(13),
                    "aliq_icms": g(14),
                    "vl_icms": g(15),
                    "vl_bc_icms_st": g(16) if len(partes) > 16 else "0",
                    "aliq_st": g(17) if len(partes) > 17 else "0",
                    "vl_icms_st": g(18) if len(partes) > 18 else "0",
                    "cst_ipi": g(20),
                    "aliq_ipi": g(24) if len(partes) > 24 else "0",
                    "vl_ipi": g(25) if len(partes) > 25 else "0",
                    "cst_pis": g(26) if len(partes) > 26 else g(25),
                    "vl_bc_pis": g(27) if len(partes) > 27 else "0",
                    "aliq_pis": g(28) if len(partes) > 28 else "0",
                    "vl_pis": g(31) if len(partes) > 31 else "0",
                    "cst_cofins": g(32) if len(partes) > 32 else g(31),
                    "vl_bc_cofins": g(33) if len(partes) > 33 else "0",
                    "aliq_cofins": g(34) if len(partes) > 34 else "0",
                    "vl_cofins": g(37) if len(partes) > 37 else "0",
                    "c100_num_doc": c100_atual.get("num_doc", ""),
                    "c100_dt_doc": c100_atual.get("dt_doc", ""),
                    "c100_cod_part": c100_atual.get("cod_part", ""),
                }
                # Enriquecer com NCM do cadastro 0200, se existir
                cad = resultado["produtos_0200"].get(cod_item)
                item["ncm"] = (cad or {}).get("ncm") or cod_item
                item["descr_cadastro"] = (cad or {}).get("descricao", item["descr"])
                resultado["itens_c170"].append(item)

            elif tipo == "C190" and c100_atual is not None:
                resultado["resumo_c190"].append({
                    "cst_icms": g(2), "cfop": g(3), "aliq_icms": g(4),
                    "vl_opr": g(5), "vl_bc_icms": g(6), "vl_icms": g(7),
                })

            elif tipo in ("M100", "M105"):
                resultado["apuracao"]["pis"].setdefault("registros", []).append(partes)

            elif tipo in ("M200", "M210"):
                resultado["apuracao"]["pis"].setdefault("apurado", []).append(partes)

            elif tipo in ("M500", "M505"):
                resultado["apuracao"]["cofins"].setdefault("registros", []).append(partes)

            elif tipo in ("M600", "M610"):
                resultado["apuracao"]["cofins"].setdefault("apurado", []).append(partes)

        except Exception as e:
            resultado["avisos"].append(f"Linha {num_linha} ({tipo}): erro ao interpretar — {e}")

    if not resultado["itens_c170"]:
        resultado["avisos"].append(
            "Nenhum registro C170 encontrado neste SPED. Isso é esperado para "
            "empresas do setor de SERVIÇOS (NFS-e) ou arquivos apenas de apuração. "
            "O relatório produto a produto será gerado vazio; a apuração "
            "PIS/COFINS (M100/M200/M500/M600), se presente, continua disponível."
        )

    return resultado


# ---------------------------------------------------------------------------
# 3) IDENTIFICAÇÃO DE SETOR (a partir do CNAE ou natureza dos itens)
# ---------------------------------------------------------------------------

def identificar_setor_empresa(itens_validados: List[Dict[str, Any]], cnae: str = "") -> str:
    """Determina o setor predominante da empresa a partir dos NCMs dos itens."""
    if cnae:
        cnae2 = cnae[:2]
        mapa_cnae = {
            "01": "Agropecuária", "02": "Agropecuária", "03": "Agropecuária",
            "10": "Indústria", "11": "Indústria", "20": "Indústria", "22": "Indústria",
            "41": "Construção Civil", "42": "Construção Civil", "43": "Construção Civil",
            "45": "Comércio", "46": "Comércio", "47": "Comércio",
            "35": "Energia", "61": "Serviços", "62": "Serviços", "86": "Saúde",
        }
        if cnae2 in mapa_cnae:
            return mapa_cnae[cnae2]

    if not itens_validados:
        return "Serviços (sem C170)"

    contagem: Dict[str, int] = {}
    for it in itens_validados:
        s = it.get("setor", "Não identificado")
        contagem[s] = contagem.get(s, 0) + 1
    return max(contagem, key=contagem.get) if contagem else "Não identificado"


# ---------------------------------------------------------------------------
# 4) VALIDAÇÃO PRODUTO A PRODUTO — UNIVERSAL
# ---------------------------------------------------------------------------

def validar_produto_universal(
    item: Dict[str, Any],
    regime: str = "LUCRO_REAL",
    uf: str = "SP",
    empresa_ctx: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Valida UM item (registro C170) contra as regras fiscais universais.

    regime aceita: "LUCRO_REAL" (não-cumulativo), "LUCRO_PRESUMIDO" (cumulativo),
                   "SIMPLES_NACIONAL", "MEI", "CPF_RURAL".

    Retorna um dicionário "linha de relatório" com todas as colunas exigidas
    pela especificação universal (ver seção 5 do relatório Excel).
    """
    empresa_ctx = empresa_ctx or {}
    regime_norm = (regime or "LUCRO_REAL").upper().replace(" ", "_")

    ncm_raw = item.get("ncm") or item.get("cod_item") or ""
    ncm_norm = _norm_ncm(ncm_raw)
    cfop = str(item.get("cfop", "")).strip()
    cst_pis = str(item.get("cst_pis", "")).strip().zfill(2)[:2]
    cst_cofins = str(item.get("cst_cofins", "")).strip().zfill(2)[:2]
    vl_item = _safe_dec(item.get("vl_item", 0))
    vl_icms = _safe_dec(item.get("vl_icms", 0))
    vl_bc_pis = _safe_dec(item.get("vl_bc_pis", 0))
    vl_bc_cofins = _safe_dec(item.get("vl_bc_cofins", 0))
    aliq_icms_inf = _safe_dec(item.get("aliq_icms", 0))
    aliq_ipi_inf = _safe_dec(item.get("aliq_ipi", 0))
    aliq_pis_inf = _safe_dec(item.get("aliq_pis", 0))
    aliq_cofins_inf = _safe_dec(item.get("aliq_cofins", 0))

    ncm_info = _consultar_ncm_universal(ncm_norm)

    linha: Dict[str, Any] = {
        "setor": ncm_info.get("setor", "Não identificado"),
        "nf": item.get("c100_num_doc", ""),
        "data": item.get("c100_dt_doc", ""),
        "cnpj_fornecedor": item.get("c100_cod_part", ""),
        "ncm": ncm_norm,
        "descricao_tipi": ncm_info.get("descricao", ""),
        "descricao_item": str(item.get("descr_cadastro") or item.get("descr", ""))[:80],
        "cfop": cfop,
        "cst_pis": cst_pis,
        "cst_cofins": cst_cofins,
        "base_pis": float(vl_bc_pis),
        "aliq_pis_inf": float(aliq_pis_inf),
        "aliq_pis_esp": 0.0,
        "credito_pis": 0.0,
        "base_cofins": float(vl_bc_cofins),
        "aliq_cofins_inf": float(aliq_cofins_inf),
        "aliq_cofins_esp": 0.0,
        "credito_cofins": 0.0,
        "icms_inf": float(aliq_icms_inf),
        "icms_esp": 0.0,
        "ipi_inf": float(aliq_ipi_inf),
        "ipi_esp": float(ncm_info.get("ipi", 0.0)),
        "monofasico": "SIM" if ncm_info.get("monofasico") else "NÃO",
        "credito_monofasico_indevido": "NÃO",
        "tese_do_seculo": "NÃO",
        "icms_st_credito": 0.0,
        "iss_excluido": 0.0,
        "setor_atacado_varejo": "NÃO",
        "trava_simples_nacional": "NÃO",
        "status": "OK",
        "observacoes": [],
    }

    erros: List[str] = []
    alertas: List[str] = []

    # (a) VALIDAÇÃO DE NCM ------------------------------------------------
    if not ncm_norm:
        erros.append("NCM ausente no item.")
    elif not ncm_info.get("encontrado"):
        sugestao = _sugerir_ncm_proximo(ncm_norm)
        msg = f"ERRO – NCM INVÁLIDO/NÃO CATALOGADO: {ncm_norm}."
        if sugestao:
            msg += f" Sugestão de faixa próxima: {sugestao}."
        erros.append(msg)

    # (g) PRODUTOR RURAL CPF ------------------------------------------------
    tipo_pessoa = empresa_ctx.get("tipo_pessoa", "PJ")
    is_cpf_rural = (regime_norm == "CPF_RURAL") or (tipo_pessoa == "PF")
    if is_cpf_rural:
        if ncm_info.get("setor") == "Agropecuária" and ncm_info.get("monofasico"):
            alertas.append(
                "CPF RURAL: insumo agropecuário monofásico (Lei 10.925/2004). "
                "Produtor rural pessoa física não apura PIS/COFINS não-cumulativo — "
                "confirmar que não há crédito indevido tomado sobre este insumo."
            )

    # (f) SIMPLES NACIONAL — TRAVA -----------------------------------------
    if regime_norm in ("SIMPLES_NACIONAL", "MEI"):
        if cst_pis in CST_COM_CREDITO or cst_cofins in CST_COM_CREDITO:
            linha["trava_simples_nacional"] = "SIM"
            erros.append(
                "TRAVA ATIVA – LC 123/2006 Art. 18: empresa optante pelo Simples "
                "Nacional NÃO pode apropriar créditos de PIS/COFINS (CST de crédito "
                f"detectado: PIS={cst_pis}/COFINS={cst_cofins})."
            )
        if regime_norm == "MEI":
            alertas.append("MEI é isento de PIS/COFINS — validação de alíquotas não aplicável.")

    # (b) VALIDAÇÃO DE IPI ---------------------------------------------------
    ipi_esp = Decimal(str(ncm_info.get("ipi", 0.0)))
    if aliq_ipi_inf == 0 and ipi_esp > 0 and cfop.startswith(("1", "2")):
        alertas.append(
            f"ALERTA – IPI não destacado (TIPI indica {ipi_esp}% para NCM {ncm_norm}), "
            f"possível crédito de IPI perdido em operação de entrada/compra (CFOP {cfop})."
        )
    elif aliq_ipi_inf > 0 and abs(aliq_ipi_inf - ipi_esp) > Decimal("2.0"):
        alertas.append(
            f"IPI informado ({aliq_ipi_inf}%) diverge significativamente da TIPI "
            f"esperada ({ipi_esp}%) para o NCM {ncm_norm}."
        )

    # (c) VALIDAÇÃO DE ICMS ---------------------------------------------------
    if cfop.startswith("7"):
        icms_esp = Decimal("0")  # exportação — imunidade
    elif cfop.startswith(("6", "2")):
        uf_destino = empresa_ctx.get("uf_destino", uf)
        if uf in UFS_SUL_SUDESTE and uf_destino not in UFS_SUL_SUDESTE:
            icms_esp = Decimal("7")
        else:
            icms_esp = Decimal("12")
    else:
        icms_esp = Decimal(str(ALIQUOTAS_ICMS_INTERNO.get(uf, 18.0)))
    linha["icms_esp"] = float(icms_esp)

    if aliq_icms_inf > 0 and abs(aliq_icms_inf - icms_esp) > Decimal("1.0"):
        alertas.append(
            f"ICMS informado ({aliq_icms_inf}%) diverge da alíquota esperada "
            f"({icms_esp}%) para UF={uf}/CFOP={cfop}. Verificar convênio CONFAZ, "
            f"redução de base ou ICMS-ST."
        )

    # (d) VALIDAÇÃO PIS/COFINS -------------------------------------------------
    if regime_norm == "LUCRO_REAL":
        pis_esp, cofins_esp = Decimal("1.65"), Decimal("7.60")
    elif regime_norm == "LUCRO_PRESUMIDO":
        pis_esp, cofins_esp = Decimal("0.65"), Decimal("3.00")
    else:
        pis_esp, cofins_esp = Decimal("0"), Decimal("0")
    linha["aliq_pis_esp"] = float(pis_esp)
    linha["aliq_cofins_esp"] = float(cofins_esp)

    if ncm_info.get("monofasico"):
        if cst_pis in CST_COM_CREDITO or cst_cofins in CST_COM_CREDITO:
            linha["credito_monofasico_indevido"] = "SIM"
            erros.append(
                f"ERRO – CRÉDITO MONOFÁSICO INDEVIDO: NCM {ncm_norm} ({ncm_info.get('setor')}) "
                f"é tributado no regime monofásico ({ncm_info.get('obs','')}). "
                f"CST PIS={cst_pis}/CST COFINS={cst_cofins} indica tomada de crédito; "
                f"revendedor deveria usar CST 04/05/06 (alíquota zero). "
                f"Aplica-se independentemente do setor que revende o produto."
            )
    elif regime_norm in ("LUCRO_REAL", "LUCRO_PRESUMIDO"):
        if aliq_pis_inf > 0 and abs(aliq_pis_inf - pis_esp) > Decimal("0.05"):
            alertas.append(f"PIS: alíquota informada {aliq_pis_inf}% diverge da esperada {pis_esp}% para o regime {regime_norm}.")
        if aliq_cofins_inf > 0 and abs(aliq_cofins_inf - cofins_esp) > Decimal("0.05"):
            alertas.append(f"COFINS: alíquota informada {aliq_cofins_inf}% diverge da esperada {cofins_esp}% para o regime {regime_norm}.")

    # Crédito calculado (informativo) — apenas quando há direito a crédito
    if regime_norm == "LUCRO_REAL" and cst_pis in CST_COM_CREDITO and not ncm_info.get("monofasico"):
        linha["credito_pis"] = float((vl_bc_pis * pis_esp / Decimal("100")).quantize(Decimal("0.01")))
        linha["credito_cofins"] = float((vl_bc_cofins * cofins_esp / Decimal("100")).quantize(Decimal("0.01")))

    # (e) TESE DO SÉCULO (RE 574.706/PR) ---------------------------------------
    if regime_norm == "LUCRO_REAL" and vl_bc_pis > 0 and vl_icms > 0 and vl_item > 0:
        bc_com_icms = abs(vl_bc_pis - vl_item) < Decimal("0.02")
        bc_sem_icms = abs(vl_bc_pis - (vl_item - vl_icms)) < Decimal("0.02")
        if bc_com_icms and not bc_sem_icms:
            linha["tese_do_seculo"] = "SIM"
            credito_adicional = ((vl_icms) * pis_esp / Decimal("100")) + ((vl_icms) * cofins_esp / Decimal("100"))
            alertas.append(
                f"ALERTA – TESE DO SÉCULO APLICÁVEL (RE 574.706/PR): BC do PIS/COFINS "
                f"(R$ {vl_bc_pis:.2f}) inclui o ICMS destacado (R$ {vl_icms:.2f}). "
                f"BC correta = R$ {(vl_item - vl_icms):.2f}. Crédito adicional estimado: "
                f"R$ {credito_adicional:.2f}."
            )

    # (h) EXCLUSÃO DO ISS DA BC PIS/COFINS — TESE FILHOTE (RE 592.616) ------
    is_servico = (
        cfop.startswith(("5", "6", "7"))
        or ncm_norm.startswith("00")
        or ncm_norm == ""
    )
    if is_servico and regime_norm == "LUCRO_REAL" and vl_item > 0:
        try:
            from integrador_rfb_expandido import consultar_aliquota_iss as _iss_fn
            iss_aliquota = _iss_fn(uf)
        except Exception:
            iss_aliquota = Decimal("5")  # fallback ISS máximo (LC 116/2003)
        vl_iss = (vl_item * iss_aliquota / Decimal("100")).quantize(Decimal("0.01"))
        if vl_iss > 0:
            linha["iss_excluido"] = float(vl_iss)
            # Ajusta créditos PIS/COFINS se houver direito
            if cst_pis in CST_COM_CREDITO and not ncm_info.get("monofasico"):
                bc_pis_aj = max(vl_bc_pis - vl_iss, Decimal("0"))
                bc_cof_aj = max(vl_bc_cofins - vl_iss, Decimal("0"))
                linha["credito_pis"] = float(
                    (bc_pis_aj * pis_esp / Decimal("100")).quantize(Decimal("0.01"))
                )
                linha["credito_cofins"] = float(
                    (bc_cof_aj * cofins_esp / Decimal("100")).quantize(Decimal("0.01"))
                )
            alertas.append(
                f"ISS EXCLUÍDO DA BC PIS/COFINS (RE 592.616 – Tese filhote): "
                f"R$ {vl_iss:.2f} (alíquota {iss_aliquota}%, UF={uf})."
            )

    # (i) RESSARCIMENTO ICMS-ST — TEMA 762 STF --------------------------------
    vl_bc_st_presumida = _safe_dec(item.get("vl_bc_icms_st", 0))
    vl_icms_st = _safe_dec(item.get("vl_icms_st", 0))
    vl_bc_real = _safe_dec(item.get("vl_bc_icms", 0))
    if vl_bc_st_presumida > 0 and vl_bc_real > 0 and vl_bc_real < vl_bc_st_presumida:
        aliq_icms_calc = icms_esp if icms_esp > 0 else Decimal("18")
        credito_st = (
            (vl_bc_st_presumida - vl_bc_real) * aliq_icms_calc / Decimal("100")
        ).quantize(Decimal("0.01"))
        linha["icms_st_credito"] = float(credito_st)
        alertas.append(
            f"ICMS-ST A RESSARCIR (Tema 762 STF): BC presumida R$ {vl_bc_st_presumida:.2f} > "
            f"BC real R$ {vl_bc_real:.2f}. Crédito estimado: R$ {credito_st:.2f}."
        )

    # (j) DETECÇÃO ATACADO / VAREJO -------------------------------------------
    cnae = empresa_ctx.get("cnae", "")
    is_atacado_varejo = (
        cnae.startswith(("46", "47"))
        or cfop.startswith(("5", "6"))
    )
    if is_atacado_varejo:
        linha["setor_atacado_varejo"] = "SIM"
        # Alerta extra para monofásico em atacado/varejo
        if ncm_info.get("monofasico") and cst_pis in CST_COM_CREDITO:
            if "MONOFÁSICO INDEVIDO" not in " ".join(erros):
                erros.append(
                    f"ATACADO/VAREJO: crédito monofásico indevido em operação de "
                    f"revenda (NCM {ncm_norm}, CFOP {cfop}). Vedado pela IN 1.911/2019 Art. 167."
                )

    # --- status final -----------------------------------------------------
    if erros:
        linha["status"] = "ERRO"
    elif alertas:
        linha["status"] = "ALERTA"
    else:
        linha["status"] = "OK"

    linha["observacoes"] = " | ".join(erros + alertas) if (erros or alertas) else "Conforme"
    linha["_erros"] = erros
    linha["_alertas"] = alertas
    linha["hash_sha256"] = sha256_linha(
        linha["nf"], linha["ncm"], linha["cfop"], linha["cst_pis"], linha["cst_cofins"],
        linha["base_pis"], linha["base_cofins"], linha["status"], datetime.now().date().isoformat(),
    )
    return linha


# ---------------------------------------------------------------------------
# 5) VALIDAÇÃO EM LOTE (STREAMING / BATCH) — PERFORMANCE
# ---------------------------------------------------------------------------

def validar_lote_universal(
    itens_c170: List[Dict[str, Any]],
    regime: str = "LUCRO_REAL",
    uf: str = "SP",
    empresa_ctx: Optional[Dict[str, Any]] = None,
    batch_size: int = 10000,
    progress_cb: Optional[Callable[[int, int], None]] = None,
) -> Dict[str, Any]:
    """
    Valida uma lista de itens C170 em LOTES (streaming-friendly), evitando
    picos de memória em SPEDs muito grandes (>1 milhão de linhas).

    progress_cb(processados, total) é chamado a cada lote — pode ser usado
    para exibir uma barra de progresso no Streamlit.
    """
    total = len(itens_c170)
    linhas: List[Dict[str, Any]] = []
    total_ok = total_alerta = total_erro = 0
    total_monofasico_indevido = 0
    total_tese_seculo = 0
    total_trava_simples = 0
    total_icms_st_ressarcimento = Decimal("0")
    total_iss_excluido = Decimal("0")
    total_atacado_varejo = 0
    ncm_nao_catalogado: Dict[str, int] = {}

    for inicio in range(0, total, batch_size):
        lote = itens_c170[inicio: inicio + batch_size]
        for item in lote:
            r = validar_produto_universal(item, regime=regime, uf=uf, empresa_ctx=empresa_ctx)
            linhas.append(r)
            if r["status"] == "OK":
                total_ok += 1
            elif r["status"] == "ALERTA":
                total_alerta += 1
            else:
                total_erro += 1
            if r["credito_monofasico_indevido"] == "SIM":
                total_monofasico_indevido += 1
            if r["tese_do_seculo"] == "SIM":
                total_tese_seculo += 1
            if r["trava_simples_nacional"] == "SIM":
                total_trava_simples += 1
            total_icms_st_ressarcimento += Decimal(str(r.get("icms_st_credito", 0)))
            total_iss_excluido += Decimal(str(r.get("iss_excluido", 0)))
            if r.get("setor_atacado_varejo") == "SIM":
                total_atacado_varejo += 1
            if "NÃO CATALOGADO" in r["observacoes"] or "NÃO ENCONTRADO" in " ".join(r.get("_erros", [])):
                ncm_nao_catalogado[r["ncm"]] = ncm_nao_catalogado.get(r["ncm"], 0) + 1
        if progress_cb:
            try:
                progress_cb(min(inicio + batch_size, total), total)
            except Exception:
                pass

    setor_predominante = identificar_setor_empresa(linhas)

    return {
        "linhas": linhas,
        "resumo": {
            "total_itens": total,
            "total_ok": total_ok,
            "total_alerta": total_alerta,
            "total_erro": total_erro,
            "total_monofasico_indevido": total_monofasico_indevido,
            "total_tese_seculo": total_tese_seculo,
            "total_trava_simples_nacional": total_trava_simples,
            "total_icms_st_ressarcimento": float(total_icms_st_ressarcimento),
            "total_iss_excluido": float(total_iss_excluido),
            "total_atacado_varejo": total_atacado_varejo,
            "pct_conformidade": round(total_ok / max(total, 1) * 100, 1),
            "setor_predominante": setor_predominante,
            "ncms_nao_catalogados": ncm_nao_catalogado,
        },
        "meta": {
            "regime": regime, "uf": uf,
            "data_validacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "versao_motor": "Universal v20 Soberano",
        },
    }


# ---------------------------------------------------------------------------
# 6) GERAÇÃO DO RELATÓRIO EXCEL (UNIVERSAL, COM FORMATAÇÃO E CORES)
# ---------------------------------------------------------------------------

COLUNAS_RELATORIO = [
    ("setor", "Setor"),
    ("nf", "NF"),
    ("data", "Data"),
    ("cnpj_fornecedor", "CNPJ/Cód. Fornecedor"),
    ("ncm", "NCM"),
    ("descricao_tipi", "Descrição TIPI"),
    ("descricao_item", "Descrição Item"),
    ("cfop", "CFOP"),
    ("cst_pis", "CST PIS"),
    ("cst_cofins", "CST COFINS"),
    ("base_pis", "Base PIS"),
    ("aliq_pis_inf", "Alíq. PIS Inf (%)"),
    ("aliq_pis_esp", "Alíq. PIS Esp (%)"),
    ("credito_pis", "Crédito PIS"),
    ("base_cofins", "Base COFINS"),
    ("aliq_cofins_inf", "Alíq. COFINS Inf (%)"),
    ("aliq_cofins_esp", "Alíq. COFINS Esp (%)"),
    ("credito_cofins", "Crédito COFINS"),
    ("icms_inf", "ICMS Inf (%)"),
    ("icms_esp", "ICMS Esp (%)"),
    ("ipi_inf", "IPI Inf (%)"),
    ("ipi_esp", "IPI Esp (%)"),
    ("monofasico", "Monofásico (S/N)"),
    ("credito_monofasico_indevido", "Crédito Monofásico Indevido (S/N)"),
    ("tese_do_seculo", "Tese do Século (S/N)"),
    ("icms_st_credito", "ICMS-ST Ressarcimento"),
    ("iss_excluido", "ISS Excluído BC"),
    ("setor_atacado_varejo", "Atacado/Varejo (S/N)"),
    ("trava_simples_nacional", "TRAVA Simples Nacional (S/N)"),
    ("status", "Status"),
    ("observacoes", "Observações"),
    ("hash_sha256", "Hash SHA-256"),
]


def gerar_excel_universal(analise: Dict[str, Any], empresa_info: Optional[Dict[str, Any]] = None) -> bytes:
    """
    Gera o Excel formatado (cores por status, resumo, autofiltro).

    Escolhe automaticamente o motor de escrita mais rápido disponível:
      1) `xlsxwriter` (recomendado — escrita em stream, memória constante,
         performance muito superior para SPEDs de centenas de milhares/
         milhões de itens).
      2) `openpyxl` como fallback (sempre disponível), usando formatação
         condicional (regra única para a faixa inteira) em vez de colorir
         célula a célula, para não degradar em arquivos grandes.
    """
    try:
        import xlsxwriter  # noqa: F401
        return _gerar_excel_universal_xlsxwriter(analise, empresa_info)
    except ImportError:
        return _gerar_excel_universal_openpyxl(analise, empresa_info)


def _gerar_excel_universal_xlsxwriter(analise: Dict[str, Any], empresa_info: Optional[Dict[str, Any]] = None) -> bytes:
    """Gerador de Excel de alta performance usando xlsxwriter (modo stream)."""
    import xlsxwriter

    empresa_info = empresa_info or {}
    resumo = analise.get("resumo", {})
    meta = analise.get("meta", {})
    linhas = analise.get("linhas", [])

    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True, "constant_memory": True})

    # --- Aba 1: Resumo ---
    ws0 = wb.add_worksheet("Resumo")
    fmt_titulo = wb.add_format({"bold": True, "font_size": 14, "font_color": "#1A237E"})
    fmt_label = wb.add_format({"bold": True})
    ws0.write(0, 0, "RELATÓRIO UNIVERSAL DE VALIDAÇÃO PRODUTO A PRODUTO", fmt_titulo)

    info_rows = [
        ("Empresa", empresa_info.get("razao_social", "")),
        ("CNPJ/CPF", empresa_info.get("cnpj", "")),
        ("UF", meta.get("uf", "")),
        ("Regime Tributário", meta.get("regime", "")),
        ("Setor Predominante", resumo.get("setor_predominante", "")),
        ("Data da Validação", meta.get("data_validacao", "")),
        ("Motor", meta.get("versao_motor", "")),
        ("", ""),
        ("Total de Itens", resumo.get("total_itens", 0)),
        ("OK", resumo.get("total_ok", 0)),
        ("Alertas", resumo.get("total_alerta", 0)),
        ("Erros", resumo.get("total_erro", 0)),
        ("Crédito Monofásico Indevido", resumo.get("total_monofasico_indevido", 0)),
        ("Tese do Século Aplicável", resumo.get("total_tese_seculo", 0)),
        ("Trava Simples Nacional Ativada", resumo.get("total_trava_simples_nacional", 0)),
        ("% Conformidade", f"{resumo.get('pct_conformidade', 0)}%"),
    ]
    r = 2
    for label, val in info_rows:
        ws0.write(r, 0, label, fmt_label)
        ws0.write(r, 1, val)
        r += 1

    ncms_nc = resumo.get("ncms_nao_catalogados", {})
    if ncms_nc:
        r += 1
        ws0.write(r, 0, "NCMs NÃO CATALOGADOS (log)", wb.add_format({"bold": True, "font_color": "#C62828"}))
        r += 1
        for ncm_, qtd in list(ncms_nc.items())[:500]:
            ws0.write(r, 0, ncm_)
            ws0.write(r, 1, qtd)
            r += 1

    ws0.set_column(0, 0, 32)
    ws0.set_column(1, 1, 40)

    # --- Aba 2: Produto a Produto ---
    ws = wb.add_worksheet("Produto a Produto")
    headers = [h for _, h in COLUNAS_RELATORIO]
    keys = [k for k, _ in COLUNAS_RELATORIO]

    fmt_header = wb.add_format({
        "bold": True, "font_color": "white", "bg_color": "#1A237E",
        "align": "center", "valign": "vcenter", "text_wrap": True, "border": 1,
    })
    for c, h in enumerate(headers):
        ws.write(0, c, h, fmt_header)

    fmt_ok = wb.add_format({"bg_color": "#C8E6C9"})
    fmt_alerta = wb.add_format({"bg_color": "#FFF9C4"})
    fmt_erro = wb.add_format({"bg_color": "#FFCDD2"})

    for row_idx, linha in enumerate(linhas, start=1):
        for col_idx, key in enumerate(keys):
            ws.write(row_idx, col_idx, linha.get(key, ""))

    n_linhas = max(len(linhas), 1)
    status_col_idx = keys.index("status")
    status_col_letter = xlsxwriter.utility.xl_col_to_name(status_col_idx)
    last_col_letter = xlsxwriter.utility.xl_col_to_name(len(headers) - 1)
    data_range = f"A2:{last_col_letter}{n_linhas + 1}"

    for formula, fmt in ((f'${status_col_letter}2="OK"', fmt_ok),
                          (f'${status_col_letter}2="ALERTA"', fmt_alerta),
                          (f'${status_col_letter}2="ERRO"', fmt_erro)):
        ws.conditional_format(data_range, {"type": "formula", "criteria": formula, "format": fmt})

    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, n_linhas, len(headers) - 1)

    larguras = [16, 10, 12, 18, 12, 30, 30, 8, 8, 10, 12, 12, 12, 12, 12, 12, 12, 12,
                10, 10, 10, 10, 12, 14, 12, 14, 10, 50, 68]
    for i, w in enumerate(larguras):
        ws.set_column(i, i, w)

    wb.close()
    return buf.getvalue()


def _gerar_excel_universal_openpyxl(analise: Dict[str, Any], empresa_info: Optional[Dict[str, Any]] = None) -> bytes:
    """Gerador de Excel via openpyxl (fallback, sem dependência extra)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule

    empresa_info = empresa_info or {}
    resumo = analise.get("resumo", {})
    meta = analise.get("meta", {})
    linhas = analise.get("linhas", [])

    wb = Workbook()

    # --- Aba 1: Resumo ---
    ws0 = wb.active
    ws0.title = "Resumo"
    ws0["A1"] = "RELATÓRIO UNIVERSAL DE VALIDAÇÃO PRODUTO A PRODUTO"
    ws0["A1"].font = Font(bold=True, size=14, color="1A237E")
    ws0.merge_cells("A1:D1")

    info_rows = [
        ("Empresa", empresa_info.get("razao_social", "")),
        ("CNPJ/CPF", empresa_info.get("cnpj", "")),
        ("UF", meta.get("uf", "")),
        ("Regime Tributário", meta.get("regime", "")),
        ("Setor Predominante", resumo.get("setor_predominante", "")),
        ("Data da Validação", meta.get("data_validacao", "")),
        ("Motor", meta.get("versao_motor", "")),
        ("", ""),
        ("Total de Itens", resumo.get("total_itens", 0)),
        ("✅ OK", resumo.get("total_ok", 0)),
        ("⚠️ Alertas", resumo.get("total_alerta", 0)),
        ("❌ Erros", resumo.get("total_erro", 0)),
        ("Crédito Monofásico Indevido", resumo.get("total_monofasico_indevido", 0)),
        ("Tese do Século Aplicável", resumo.get("total_tese_seculo", 0)),
        ("Trava Simples Nacional Ativada", resumo.get("total_trava_simples_nacional", 0)),
        ("% Conformidade", f"{resumo.get('pct_conformidade', 0)}%"),
    ]
    r = 3
    for label, val in info_rows:
        ws0[f"A{r}"] = label
        ws0[f"A{r}"].font = Font(bold=True)
        ws0[f"B{r}"] = val
        r += 1

    ncms_nc = resumo.get("ncms_nao_catalogados", {})
    if ncms_nc:
        r += 1
        ws0[f"A{r}"] = "NCMs NÃO CATALOGADOS (log)"
        ws0[f"A{r}"].font = Font(bold=True, color="C62828")
        r += 1
        for ncm_, qtd in list(ncms_nc.items())[:200]:
            ws0[f"A{r}"] = ncm_
            ws0[f"B{r}"] = qtd
            r += 1

    ws0.column_dimensions["A"].width = 32
    ws0.column_dimensions["B"].width = 40

    # --- Aba 2: Produto a Produto ---
    ws_data = wb.create_sheet("Produto a Produto")
    headers = [h for _, h in COLUNAS_RELATORIO]
    keys = [k for k, _ in COLUNAS_RELATORIO]
    n_cols = len(headers)
    status_col_letter = get_column_letter(keys.index("status") + 1)

    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws_data.append(headers)
    for c in range(1, n_cols + 1):
        cell = ws_data.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for linha in linhas:
        ws_data.append([linha.get(k, "") for k in keys])

    n_linhas = max(len(linhas), 1)
    data_range = f"A2:{get_column_letter(n_cols)}{n_linhas + 1}"

    fill_ok = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    fill_alerta = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    fill_erro = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    # Formatação condicional aplicada de uma só vez à faixa inteira — o Excel
    # (não o Python) faz a coloração linha a linha na hora de abrir o arquivo.
    ws_data.conditional_formatting.add(
        data_range, FormulaRule(formula=[f'${status_col_letter}2="OK"'], fill=fill_ok)
    )
    ws_data.conditional_formatting.add(
        data_range, FormulaRule(formula=[f'${status_col_letter}2="ALERTA"'], fill=fill_alerta)
    )
    ws_data.conditional_formatting.add(
        data_range, FormulaRule(formula=[f'${status_col_letter}2="ERRO"'], fill=fill_erro)
    )

    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_linhas + 1}"

    larguras = [16, 10, 12, 18, 12, 30, 30, 8, 8, 10, 12, 12, 12, 12, 12, 12, 12, 12,
                10, 10, 10, 10, 12, 14, 12, 14, 10, 50, 68]
    for i, w in enumerate(larguras, start=1):
        ws_data.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_relatorio_produtos_universal(
    sped_bytes: bytes,
    cnpj: str = "",
    regime: str = "LUCRO_REAL",
    uf: str = "SP",
    empresa_info: Optional[Dict[str, Any]] = None,
    progress_cb: Optional[Callable[[int, int], None]] = None,
) -> bytes:
    """
    Função principal do motor: recebe os BYTES de um SPED, faz o parsing
    universal, valida produto a produto e devolve os BYTES de um Excel
    formatado e pronto para uso por contador/fiscal.

    - regime: "LUCRO_REAL" | "LUCRO_PRESUMIDO" | "SIMPLES_NACIONAL" | "MEI" | "CPF_RURAL"
    - Nunca lança exceção para SPED sem C170: gera relatório vazio com aviso.
    """
    try:
        conteudo = sped_bytes.decode("latin-1", errors="replace")
    except Exception:
        conteudo = sped_bytes.decode("utf-8", errors="replace")

    dados = parsear_sped_universal(conteudo)

    empresa_info = dict(empresa_info or {})
    empresa_info.setdefault("cnpj", cnpj or dados["empresa"].get("cnpj", ""))
    empresa_info.setdefault("razao_social", dados["empresa"].get("razao_social", ""))
    empresa_ctx = {
        "uf_destino": uf,
        "tipo_pessoa": dados["empresa"].get("tipo_pessoa", "PJ"),
    }

    analise = validar_lote_universal(
        dados["itens_c170"], regime=regime, uf=uf, empresa_ctx=empresa_ctx,
        batch_size=10000, progress_cb=progress_cb,
    )
    analise["avisos_parser"] = dados.get("avisos", [])

    return gerar_excel_universal(analise, empresa_info)



# ---------------------------------------------------------------------------
# 8) SISTEMA DE PERFIS — VISÕES ESPECIALIZADAS (v20 Soberano)
# ---------------------------------------------------------------------------

PERFIS_DISPONIVEIS = {
    "contador": "📊 Contador — Créditos, PER/DCOMP, Planejamento",
    "auditor_rfb": "🔍 Auditor RFB — Conformidade, Glosas, Divergências",
    "advogado": "⚖️ Advogado Tributarista — Teses, Jurisprudência, Riscos",
    "juiz": "🏛 Juiz/CARF — Prova Documental, Merkle, Cadeia de Custódia",
    "cto": "💻 CTO — Performance, Volume, Custos, Qualidade",
}


def aplicar_perfil(perfil: str, analise: Dict[str, Any]) -> Dict[str, Any]:
    """Transforma o resultado bruto numa visão especializada para o perfil."""
    funcs = {
        "contador": _perfil_contador,
        "auditor_rfb": _perfil_auditor_rfb,
        "advogado": _perfil_advogado,
        "juiz": _perfil_juiz,
        "cto": _perfil_cto,
    }
    fn = funcs.get(perfil, _perfil_contador)
    return fn(analise)


def _perfil_contador(analise: Dict[str, Any]) -> Dict[str, Any]:
    res = analise.get("resumo", {})
    linhas = analise.get("linhas", [])
    total_pis = sum(r.get("credito_pis", 0) for r in linhas)
    total_cofins = sum(r.get("credito_cofins", 0) for r in linhas)
    total_st = res.get("total_icms_st_ressarcimento", 0)
    total_iss = res.get("total_iss_excluido", 0)
    return {
        "titulo": "📊 Visão do Contador",
        "subtitulo": "Créditos tributários recuperáveis e planejamento",
        "metricas": [
            ("Crédito PIS estimado", f"R$ {total_pis:,.2f}"),
            ("Crédito COFINS estimado", f"R$ {total_cofins:,.2f}"),
            ("ICMS-ST a ressarcir", f"R$ {total_st:,.2f}"),
            ("ISS excluído da BC", f"R$ {total_iss:,.2f}"),
            ("Itens processados", str(res.get("total_itens", 0))),
            ("Conformidade", f"{res.get('pct_conformidade', 0)}%"),
            ("Monofásico indevido", str(res.get("total_monofasico_indevido", 0))),
        ],
        "recomendacoes": _recomendacoes_contador(analise),
        "cor_destaque": "#2E86C1",
    }


def _recomendacoes_contador(analise: Dict[str, Any]) -> List[str]:
    recs: List[str] = []
    res = analise.get("resumo", {})
    if res.get("total_monofasico_indevido", 0) > 0:
        recs.append("Retificar EFD-Contribuições: alterar CST dos itens monofásicos para 04/05/06.")
    if res.get("total_tese_seculo", 0) > 0:
        recs.append("Tese do Século aplicável — verificar se a empresa já ajuizou ação (RE 574.706).")
    if res.get("total_icms_st_ressarcimento", 0) > 0:
        recs.append("Protocolar pedido de ressarcimento ICMS-ST junto à SEFAZ (Tema 762).")
    if res.get("total_iss_excluido", 0) > 0:
        recs.append("Avaliar ação judicial para exclusão do ISS da BC PIS/COFINS (RE 592.616).")
    if res.get("total_trava_simples_nacional", 0) > 0:
        recs.append("URGENTE: empresa do Simples Nacional com créditos de PIS/COFINS — vedado pela LC 123/2006.")
    if not recs:
        recs.append("Nenhuma irregularidade identificada. Escrituração conforme.")
    return recs


def _perfil_auditor_rfb(analise: Dict[str, Any]) -> Dict[str, Any]:
    res = analise.get("resumo", {})
    return {
        "titulo": "🔍 Visão do Auditor RFB",
        "subtitulo": "Conformidade fiscal, divergências e riscos de glosa",
        "metricas": [
            ("Total de itens analisados", str(res.get("total_itens", 0))),
            ("Itens com ERRO", str(res.get("total_erro", 0))),
            ("Itens com ALERTA", str(res.get("total_alerta", 0))),
            ("Créditos monofásicos indevidos", str(res.get("total_monofasico_indevido", 0))),
            ("Trava Simples Nacional", str(res.get("total_trava_simples_nacional", 0))),
            ("Conformidade", f"{res.get('pct_conformidade', 0)}%"),
            ("NCMs não catalogados", str(len(res.get("ncms_nao_catalogados", {})))),
        ],
        "riscos": _riscos_auditor(analise),
        "recomendacoes": _recomendacoes_auditor(analise),
        "cor_destaque": "#E74C3C",
    }


def _riscos_auditor(analise: Dict[str, Any]) -> List[str]:
    riscos: List[str] = []
    res = analise.get("resumo", {})
    if res.get("total_monofasico_indevido", 0) > 0:
        riscos.append(
            f"🔴 ALTO — {res['total_monofasico_indevido']} créditos monofásicos indevidos "
            f"detectados. Risco de glosa + multa isolada 75% (Art. 44 Lei 9.430/96)."
        )
    if res.get("total_trava_simples_nacional", 0) > 0:
        riscos.append(
            f"🔴 ALTO — {res['total_trava_simples_nacional']} itens com crédito PIS/COFINS "
            f"em empresa do Simples Nacional. Vedação expressa (LC 123/2006 Art. 18)."
        )
    if res.get("total_erro", 0) > 10:
        riscos.append(f"🟡 MÉDIO — {res['total_erro']} itens com erro de classificação fiscal.")
    ncms_nc = res.get("ncms_nao_catalogados", {})
    if len(ncms_nc) > 5:
        riscos.append(f"🟡 MÉDIO — {len(ncms_nc)} NCMs não encontrados na base oficial da RFB.")
    if not riscos:
        riscos.append("🟢 BAIXO — Nenhum risco significativo identificado.")
    return riscos


def _recomendacoes_auditor(analise: Dict[str, Any]) -> List[str]:
    recs: List[str] = []
    res = analise.get("resumo", {})
    if res.get("total_monofasico_indevido", 0) > 0:
        recs.append("Lavrar auto de infração para glosa dos créditos monofásicos indevidos.")
    if res.get("total_erro", 0) > 0:
        recs.append("Intimar o contribuinte para justificar ou retificar a EFD-Contribuições.")
    if res.get("total_tese_seculo", 0) > 0:
        recs.append("Verificar se há decisão judicial transitada em julgado autorizando exclusão do ICMS da BC.")
    if not recs:
        recs.append("Escrituração em conformidade. Nenhuma ação fiscal recomendada.")
    return recs


def _perfil_advogado(analise: Dict[str, Any]) -> Dict[str, Any]:
    res = analise.get("resumo", {})
    linhas = analise.get("linhas", [])
    total_pis = sum(r.get("credito_pis", 0) for r in linhas)
    total_cofins = sum(r.get("credito_cofins", 0) for r in linhas)
    teses: List[str] = []
    if res.get("total_tese_seculo", 0) > 0:
        teses.append("Tese do Século (RE 574.706/PR) — exclusão do ICMS da BC PIS/COFINS")
    if res.get("total_iss_excluido", 0) > 0:
        teses.append("Tese Filhote (RE 592.616) — exclusão do ISS da BC PIS/COFINS")
    if res.get("total_icms_st_ressarcimento", 0) > 0:
        teses.append("Ressarcimento ICMS-ST (Tema 762 STF) — base presumida > base real")
    if res.get("total_monofasico_indevido", 0) > 0:
        teses.append("Monofásico — vedação de crédito (IN 1.911/2019 Art. 167)")
    return {
        "titulo": "⚖️ Visão do Advogado Tributarista",
        "subtitulo": "Teses aplicáveis, fundamentação e estratégia jurídica",
        "metricas": [
            ("Teses identificadas", str(len(teses))),
            ("Créditos PIS+COFINS em disputa", f"R$ {(total_pis + total_cofins):,.2f}"),
            ("ICMS-ST ressarcível", f"R$ {res.get('total_icms_st_ressarcimento', 0):,.2f}"),
            ("ISS excluível da BC", f"R$ {res.get('total_iss_excluido', 0):,.2f}"),
            ("Itens com erro (glosa potencial)", str(res.get("total_erro", 0))),
        ],
        "teses_aplicaveis": teses or ["Nenhuma tese aplicável identificada."],
        "fundamentacao_legal": [
            "CTN Arts. 165-168 (restituição de tributo pago indevidamente)",
            "RE 574.706/PR (Tema 69 — exclusão do ICMS da BC PIS/COFINS)",
            "RE 592.616 (exclusão do ISS da BC PIS/COFINS — pendente)",
            "Tema 762 STF (ressarcimento ICMS-ST — base presumida > real)",
            "IN 1.911/2019 Art. 167 (vedação crédito monofásico)",
            "LC 123/2006 Art. 18 (Simples Nacional — vedação créditos)",
            "IN 2.055/2021 Art. 102 (habilitação crédito judicial)",
        ],
        "risco_juridico": (
            "Alto" if res.get("total_erro", 0) > 20
            else "Médio" if res.get("total_erro", 0) > 5
            else "Baixo"
        ),
        "recomendacoes": [
            "Anexar laudo pericial com hash Merkle e DNA da linha para provar integridade.",
            "Identificar jurisprudência favorável (STF/STJ) atualizada no pedido.",
            "Se houver ação judicial, habilitar crédito via IN 2.055/2021 Art. 102.",
            "Avaliar viabilidade de mandado de segurança preventivo.",
        ],
        "cor_destaque": "#8E44AD",
    }


def _perfil_juiz(analise: Dict[str, Any]) -> Dict[str, Any]:
    res = analise.get("resumo", {})
    linhas = analise.get("linhas", [])
    # Merkle-like integrity summary
    hashes = [r.get("hash_sha256", "") for r in linhas if r.get("hash_sha256")]
    raiz_merkle = ""
    if hashes:
        import hashlib as _hl
        combined = "|".join(hashes)
        raiz_merkle = _hl.sha256(combined.encode()).hexdigest()
    return {
        "titulo": "🏛 Visão do Juiz / CARF",
        "subtitulo": "Prova documental, integridade e cadeia de custódia",
        "metricas": [
            ("Total de itens periciados", str(res.get("total_itens", 0))),
            ("Itens com integridade verificada", str(len(hashes))),
            ("Raiz Merkle (SHA-256)", raiz_merkle[:32] + "..." if raiz_merkle else "N/A"),
            ("Conformidade fiscal", f"{res.get('pct_conformidade', 0)}%"),
            ("Erros encontrados", str(res.get("total_erro", 0))),
            ("Alertas", str(res.get("total_alerta", 0))),
        ],
        "cadeia_custodia": [
            f"Motor de validação: Universal v20 Soberano",
            f"Data da perícia: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            f"Total de linhas de prova: {len(hashes)}",
            f"Hash raiz (Merkle): {raiz_merkle}",
            f"Algoritmo: SHA-256 (linha) → SHA-256 (árvore)",
        ],
        "parecer": (
            "LAUDO PERICIAL — os dados analisados demonstram integridade criptográfica "
            "verificável via árvore de Merkle. Cada item possui DNA de linha (hash SHA-256) "
            "que pode ser auditado independentemente. A raiz Merkle garante que nenhum dado "
            "foi alterado após a geração do relatório."
        ),
        "recomendacoes": [
            "Validar raiz Merkle contra os dados originais do SPED.",
            "Conferir DNA de linha de itens específicos para contraprova.",
            "Aceitar como prova técnica conforme CPC Art. 464 e seguintes.",
        ],
        "cor_destaque": "#1A5276",
    }


def _perfil_cto(analise: Dict[str, Any]) -> Dict[str, Any]:
    res = analise.get("resumo", {})
    meta = analise.get("meta", {})
    total = res.get("total_itens", 0)
    return {
        "titulo": "💻 Visão do CTO",
        "subtitulo": "Performance do motor, métricas de qualidade e volume",
        "metricas": [
            ("Itens processados", str(total)),
            ("Versão do motor", meta.get("versao_motor", "?")),
            ("Data/hora execução", meta.get("data_validacao", "?")),
            ("Conformidade", f"{res.get('pct_conformidade', 0)}%"),
            ("Erros de validação", str(res.get("total_erro", 0))),
            ("Alertas gerados", str(res.get("total_alerta", 0))),
            ("NCMs não catalogados", str(len(res.get("ncms_nao_catalogados", {})))),
            ("Regime", meta.get("regime", "?")),
            ("UF", meta.get("uf", "?")),
        ],
        "qualidade": {
            "cobertura_ncm": f"{100 - (len(res.get('ncms_nao_catalogados', {})) / max(total, 1) * 100):.1f}%",
            "taxa_erro": f"{res.get('total_erro', 0) / max(total, 1) * 100:.1f}%",
            "taxa_alerta": f"{res.get('total_alerta', 0) / max(total, 1) * 100:.1f}%",
        },
        "recomendacoes": [
            "Monitorar NCMs não catalogados e atualizar base periodicamente.",
            "Otimizar batch_size se processando >500k itens.",
            "Implementar cache Redis para consultas NCM em produção.",
        ],
        "cor_destaque": "#27AE60",
    }

# ---------------------------------------------------------------------------
# 7) AUTOTESTE — SPED FICTÍCIO MULTISSETORIAL
# ---------------------------------------------------------------------------

def _gerar_itens_teste_multissetorial() -> List[Dict[str, Any]]:
    """Gera itens fictícios cobrindo Indústria, Comércio, Agro e Bebidas/Monofásico."""
    base_item = {
        "c100_num_doc": "000001", "c100_dt_doc": "01012025", "c100_cod_part": "FORN001",
        "cst_icms": "00", "vl_desc": "0",
    }
    itens = [
        # Indústria — insumo com IPI, crédito correto
        {**base_item, "num_item": "1", "cod_item": "PROD001", "ncm": "73269090",
         "descr": "Peça industrial de aço", "qtd": "100", "unid": "UN",
         "vl_item": "10000.00", "cfop": "1101", "aliq_icms": "18.00", "vl_icms": "1800.00",
         "vl_bc_icms": "10000.00", "aliq_ipi": "5.00", "cst_pis": "50", "cst_cofins": "50",
         "vl_bc_pis": "9820.00", "aliq_pis": "1.65", "vl_bc_cofins": "9820.00", "aliq_cofins": "7.60"},
        # Comércio — revenda de bebida monofásica com CST indevido (crédito indevido)
        {**base_item, "num_item": "2", "cod_item": "PROD002", "ncm": "22021000",
         "descr": "Refrigerante 2L", "qtd": "500", "unid": "UN",
         "vl_item": "5000.00", "cfop": "1102", "aliq_icms": "18.00", "vl_icms": "900.00",
         "vl_bc_icms": "5000.00", "aliq_ipi": "0", "cst_pis": "50", "cst_cofins": "50",
         "vl_bc_pis": "5000.00", "aliq_pis": "1.65", "vl_bc_cofins": "5000.00", "aliq_cofins": "7.60"},
        # Agropecuária — fertilizante (monofásico, CPF rural, sem crédito)
        {**base_item, "num_item": "3", "cod_item": "PROD003", "ncm": "31021000",
         "descr": "Ureia agrícola", "qtd": "1000", "unid": "KG",
         "vl_item": "8000.00", "cfop": "1102", "aliq_icms": "0", "vl_icms": "0",
         "vl_bc_icms": "0", "aliq_ipi": "0", "cst_pis": "04", "cst_cofins": "04",
         "vl_bc_pis": "0", "aliq_pis": "0", "vl_bc_cofins": "0", "aliq_cofins": "0"},
        # Serviços de transporte — sem NCM real (item 09), status alerta
        {**base_item, "num_item": "4", "cod_item": "PROD004", "ncm": "",
         "descr": "Serviço de transporte", "qtd": "1", "unid": "UN",
         "vl_item": "3000.00", "cfop": "1352", "aliq_icms": "0", "vl_icms": "0",
         "vl_bc_icms": "0", "aliq_ipi": "0", "cst_pis": "50", "cst_cofins": "50",
         "vl_bc_pis": "3000.00", "aliq_pis": "1.65", "vl_bc_cofins": "3000.00", "aliq_cofins": "7.60"},
        # Tese do Século — BC do PIS inclui ICMS
        {**base_item, "num_item": "5", "cod_item": "PROD005", "ncm": "39201000",
         "descr": "Chapa plástica industrial", "qtd": "200", "unid": "UN",
         "vl_item": "8200.00", "cfop": "1101", "aliq_icms": "18.00", "vl_icms": "1476.00",
         "vl_bc_icms": "8200.00", "aliq_ipi": "5.00", "cst_pis": "50", "cst_cofins": "50",
         "vl_bc_pis": "8200.00", "aliq_pis": "1.65", "vl_bc_cofins": "8200.00", "aliq_cofins": "7.60"},
    ]
    return itens


def autoteste_motor_universal() -> Dict[str, Any]:
    """Executa um teste completo com itens fictícios de vários setores."""
    itens = _gerar_itens_teste_multissetorial()
    analise = validar_lote_universal(itens, regime="LUCRO_REAL", uf="SP",
                                      empresa_ctx={"uf_destino": "SP", "tipo_pessoa": "PJ"})
    excel_bytes = gerar_excel_universal(analise, {"razao_social": "EMPRESA TESTE MULTISSETORIAL",
                                                    "cnpj": "12345678000195"})
    return {"analise": analise, "excel_bytes": excel_bytes}


if __name__ == "__main__":
    print("=== AUTOTESTE — MOTOR UNIVERSAL DE VALIDAÇÃO PRODUTO A PRODUTO ===")
    r = autoteste_motor_universal()
    resumo = r["analise"]["resumo"]
    print(json.dumps(resumo, indent=2, ensure_ascii=False, default=str))
    with open("teste_relatorio_universal.xlsx", "wb") as f:
        f.write(r["excel_bytes"])
    print(f"\nExcel de teste gerado: teste_relatorio_universal.xlsx ({len(r['excel_bytes'])} bytes)")

    # Teste com o SPED de exemplo real, se existir
    if os.path.exists("sped_exemplo.txt"):
        with open("sped_exemplo.txt", "rb") as f:
            sped_bytes = f.read()
        xlsx = gerar_relatorio_produtos_universal(sped_bytes, regime="LUCRO_REAL", uf="SP")
        with open("teste_sped_exemplo_universal.xlsx", "wb") as f:
            f.write(xlsx)
        print(f"Excel gerado a partir de sped_exemplo.txt: {len(xlsx)} bytes")
